import json
import os
import requests
from openpyxl.utils import get_column_letter

import ExcelFileUtils
import Constants

# apk打包数据的开始单元
SHEET_NUM_PACKAGE_INFO_START = 'A4'
# params参数json的所在单元
SHEET_NUM_PARAMS_JSON = 'B2'


def parse(file):
    sheet = ExcelFileUtils.get_sheet(file, 'Sheet1')
    highestRow = sheet.max_row
    highestColumn = sheet.max_column
    c = get_column_letter(highestColumn)
    c = c + str(highestRow)
    result = {}
    # 获取打包配置
    raw = sheet[SHEET_NUM_PARAMS_JSON].value
    if raw is not None and len(raw) > 0:
        params = json.loads(raw)
        if params is not None and len(params) > 0:
            result[Constants.KEY_PARAMS] = params
    # 获取每个包的配置信息
    infoList = []
    for rowOfCellObjects in sheet[SHEET_NUM_PACKAGE_INFO_START:c]:
        element = {}
        for cellObj in rowOfCellObjects:
            coordinate = cellObj.coordinate
            if coordinate.startswith("A"):
                # 处理app名称
                name = cellObj.value
                if name is None or name.isspace():
                    break
                element[Constants.KEY_INFO_APP_NAME] = name
            elif coordinate.startswith("B"):
                # 处理图片url
                element[Constants.KEY_INFO_APP_ICON_URL] = cellObj.value
            elif coordinate.startswith("C"):
                # 处理渠道前缀
                element[Constants.KEY_INFO_CHANNEL_PREFIXES] = cellObj.value
            elif coordinate.startswith("D"):
                # 处理渠道
                raw = cellObj.value
                element[Constants.KEY_INFO_CHANNEL] = raw.split(r',')
        if len(element) > 0:
            infoList.append(element)
            result[Constants.KEY_INFO_LIST] = infoList
    return result


def init_apk_json(info_list):
    for i in info_list:
        iconUrl = i[Constants.KEY_INFO_APP_ICON_URL]
        request_save_icon(iconUrl)


# 请求下载图片并保存
def request_save_icon(icon_url):
    if not os.path.exists(Constants.PATH_MATERIAL_ICON):
        os.mkdir(Constants.PATH_MATERIAL_ICON)
    imageName = parse_image_name(icon_url)
    imagePath = os.path.join(Constants.PATH_MATERIAL_ICON, imageName)
    # 判断本地是否含有同名文件
    if not os.path.exists(imagePath):
        print('下载图片 Url=' + icon_url)
        # Download the image.
        res = requests.get(icon_url)
        res.raise_for_status()
        # Save the image to
        imageFile = open(imagePath, 'wb')
        for chunk in res.iter_content(100000):
            imageFile.write(chunk)
        imageFile.close()


# 解析图片名称
def parse_image_name(icon_url):
    imageName = os.path.basename(icon_url)
    if r'?' in imageName:
        imageName = imageName.partition(r'?')[0]
    return imageName
